"""
Pilot xlsx build — 2 sheets from research/04 + research/12.

Sheet `4_Deleverage` — implements research/04 §8b formula spec:
   G_gap     = NGDP_yoy − LT_Rate
   pi_total  = dM0_pctGDP + dCB_pctGDP   (4Q deltas, both as % of GDP)
   regime    = IFS truth table (UGLY_INFLATIONARY / BEAUTIFUL /
               UGLY_DEFLATIONARY / TRANSITIONAL)
   beautiful = AND-band gate

Inputs are research/04 §7 worked-case rows (US 1930-32, US 1933-37,
Japan 1990+) — Dalio-cited values, byte-exact.

Sheet `12_Stress` — implements research/12 §5 + §7 Table 7.1:
   Shock matrix S (5×4) — DERIVED per archetype anchors.
   Weights w = (0.30, 0.40, 0.15, 0.075, 0.075) — Robbins (2014).
   Contributions C[i,e] = w[i] · S[i,e] (Excel SUMPRODUCT for column sums).

Output: pilot/dalio_model.xlsx
Run:    python pilot/build_xlsx.py
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PILOT_DIR = Path(__file__).resolve().parent
OUT = PILOT_DIR / "dalio_model.xlsx"

# ----------------------------------------------------------------------------
# Locked palette (12 tokens) — research/01 §8c
# ----------------------------------------------------------------------------
PAL = {
    "bg_page":    "FF0B0B0B",
    "bg_panel":   "FF141414",
    "bg_inset":   "FF080808",
    "bg_tooltip": "FF1C1C1C",
    "border":     "FF262626",
    "grid":       "FF1C1C1C",
    "text_1":     "FFF5F5F5",
    "text_2":     "FFA3A3A3",
    "text_3":     "FF6B7280",
    "good":       "FF00D08C",
    "trend":      "FF7FFFD4",
    "warn":       "FFD4A373",
    "bad":        "FFE5484D",
}

# Common styles
F_HEADER = Font(name="Inter", size=11, bold=True, color=PAL["text_1"])
F_BODY   = Font(name="Inter", size=10, color=PAL["text_1"])
F_NOTE   = Font(name="Inter", size=9,  color=PAL["text_2"], italic=True)
F_LABEL  = Font(name="Inter", size=10, color=PAL["text_2"], bold=True)

FILL_PAGE  = PatternFill(start_color=PAL["bg_page"],  end_color=PAL["bg_page"],  fill_type="solid")
FILL_PANEL = PatternFill(start_color=PAL["bg_panel"], end_color=PAL["bg_panel"], fill_type="solid")
FILL_INSET = PatternFill(start_color=PAL["bg_inset"], end_color=PAL["bg_inset"], fill_type="solid")
FILL_HEAD  = PatternFill(start_color=PAL["bg_tooltip"], end_color=PAL["bg_tooltip"], fill_type="solid")

BORDER_THIN = Border(
    left=Side(style="thin", color=PAL["border"]),
    right=Side(style="thin", color=PAL["border"]),
    top=Side(style="thin", color=PAL["border"]),
    bottom=Side(style="thin", color=PAL["border"]),
)

ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")


def style_cell(c, font=F_BODY, fill=FILL_PANEL, border=BORDER_THIN, align=ALIGN_RIGHT):
    c.font = font
    c.fill = fill
    c.border = border
    c.alignment = align


def header_cell(c):
    style_cell(c, font=F_HEADER, fill=FILL_HEAD, align=ALIGN_CENTER)


def label_cell(c):
    style_cell(c, font=F_LABEL, fill=FILL_PANEL, align=ALIGN_LEFT)


def note_cell(c):
    style_cell(c, font=F_NOTE, fill=FILL_INSET, align=ALIGN_LEFT)


# ----------------------------------------------------------------------------
# Sheet 4_Deleverage — research/04 worked cases + §8b formulas
# ----------------------------------------------------------------------------

def build_deleverage_sheet(ws):
    # Title row
    ws["A1"] = "1.4 Deleveragings — four-lever decomposition + regime gate"
    ws["A1"].font = Font(name="Inter", size=14, bold=True, color=PAL["text_1"])
    ws["A1"].fill = FILL_PANEL
    ws.merge_cells("A1:J1")

    ws["A2"] = "Source: research/04_deleveragings.md §7 (worked values) + §8b (formulas). Inputs Dalio-cited; outputs computed."
    ws["A2"].font = F_NOTE
    ws["A2"].fill = FILL_INSET
    ws.merge_cells("A2:J2")

    # Header row 4 — input columns
    headers = [
        "case",
        "NGDP_yoy (%)",
        "LT_Rate (%)",
        "DebtGDP_t (%)",
        "DebtGDP_t-4 (%)",
        "M0_pctGDP_t (%)",
        "M0_pctGDP_t-4 (%)",
        "CB_pctGDP_t (%)",
        "CB_pctGDP_t-4 (%)",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        header_cell(c)

    # Output columns (computed)
    out_headers = [
        "G_gap (pp)",
        "ΔDebtGDP (pp)",
        "pi_total (pp)",
        "regime",
        "beautiful",
    ]
    for j, h in enumerate(out_headers, start=10):
        c = ws.cell(row=4, column=j, value=h)
        header_cell(c)

    # Data rows — Dalio §7 anchors. Where t-4 lag is not Dalio-published,
    # use stylized values that produce the §7-stated G/ΔD/π totals.
    cases = [
        # (label, NGDP_yoy, LT_Rate, DebtGDP_t, DebtGDP_t-4, M0_t, M0_t-4, CB_t, CB_t-4)
        ("US 1930-32",  -17.0, 3.4, 252.0, 155.0, 0.4, 0.0, 0.4, 0.0),
        ("US 1933-37",  9.2, 2.9, 235.0, 252.0, 1.7, 0.0, 0.3, 0.0),
        ("Japan 1990+", 0.6, 2.6, 498.0, 403.0, 0.7, 0.0, 0.1, 0.0),
    ]
    for r, row in enumerate(cases, start=5):
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=val)
            if col == 1:
                label_cell(c)
            else:
                style_cell(c)

        # Formulas — §8b verbatim
        # J: G_gap = B - C
        c = ws.cell(row=r, column=10, value=f"=B{r}-C{r}")
        style_cell(c)
        # K: ΔDebtGDP = D - E
        c = ws.cell(row=r, column=11, value=f"=D{r}-E{r}")
        style_cell(c)
        # L: pi_total = (F-G) + (H-I)
        c = ws.cell(row=r, column=12, value=f"=(F{r}-G{r})+(H{r}-I{r})")
        style_cell(c)
        # M: regime via §8b IFS
        regime_formula = (
            f'=IFS('
            f'AND(J{r}>0,L{r}>0,J{r}>0,L{r}>0,L{r}<0.5),"UGLY_INFLATIONARY",'   # placeholder (CPI/FX gates not in this minimal sheet)
            f'AND(J{r}>0,K{r}<0,L{r}>=0.5,L{r}<=4),"BEAUTIFUL",'
            f'AND(J{r}<0,K{r}>0),"UGLY_DEFLATIONARY",'
            f'TRUE,"TRANSITIONAL")'
        )
        c = ws.cell(row=r, column=13, value=regime_formula)
        style_cell(c, align=ALIGN_CENTER)
        # N: beautiful gate
        beautiful_formula = (
            f"=IF(AND(J{r}>=0,J{r}<=3,K{r}<0,L{r}>=0.5,L{r}<=4),1,0)"
        )
        c = ws.cell(row=r, column=14, value=beautiful_formula)
        style_cell(c)

    # Footer note
    ws.cell(row=9, column=1, value="DERIVED notes:")
    ws.cell(row=9, column=1).font = F_LABEL
    ws.cell(row=9, column=1).fill = FILL_PANEL
    ws.merge_cells("A9:N9")

    notes = [
        "• G_gap, ΔDebtGDP, pi_total formulas verbatim from research/04 §8b.",
        "• regime IFS shown is the deleveragings-only path; full version has CPI_yoy and FX_Gold gates (need additional inputs).",
        "• t-4 lag values for M0/CB stylized (Dalio §7 reports level-only). G_gap and pi_total totals match §7 worked text.",
        "• Open in Excel → values auto-compute. Formula bar shows §8b expressions verbatim.",
    ]
    for k, n in enumerate(notes, start=10):
        c = ws.cell(row=k, column=1, value=n)
        note_cell(c)
        ws.merge_cells(start_row=k, start_column=1, end_row=k, end_column=14)

    # Column widths
    widths = [14, 14, 14, 16, 16, 16, 16, 16, 16, 14, 14, 14, 22, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.sheet_view.showGridLines = False

    # Sheet-level fill
    ws.sheet_properties.tabColor = PAL["good"][2:]


# ----------------------------------------------------------------------------
# Sheet 12_Stress — research/12 §5 + §7 Table 7.1
# ----------------------------------------------------------------------------

def build_stress_sheet(ws):
    ws["A1"] = "2.5 Stress Testing — All-Weather portfolio under Dalio archetypes"
    ws["A1"].font = Font(name="Inter", size=14, bold=True, color=PAL["text_1"])
    ws["A1"].fill = FILL_PANEL
    ws.merge_cells("A1:G1")

    ws["A2"] = "Source: research/12 §5 (shock matrix S) + §7 Table 7.1 (contributions). Per-archetype return = SUMPRODUCT(w, column of S)."
    ws["A2"].font = F_NOTE
    ws["A2"].fill = FILL_INSET
    ws.merge_cells("A2:G2")

    # Section 1: shock matrix S
    ws["A4"] = "Step 1 — Archetype shock matrix S (5 sleeves × 4 archetypes)"
    ws["A4"].font = F_LABEL
    ws["A4"].fill = FILL_PANEL
    ws.merge_cells("A4:G4")

    # Header row 5
    cols = ["sleeve", "weight w_i", "Defl.", "Infl.", "Stag.", "Refl."]
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=5, column=i, value=h)
        header_cell(c)

    # Sleeves + weights (research/12 §5 + §7)
    sleeves = [
        ("SPX",         0.300, -0.50, -0.30, -0.37, +0.25),
        ("Long Tsy",    0.400, +0.20, -0.50, -0.05, +0.05),
        ("Int Tsy",     0.150, +0.10, -0.40, +0.02, +0.03),
        ("Gold",        0.075,  0.00, +0.80, +1.00, +0.10),
        ("Commodities", 0.075, -0.35, +0.40, +0.30, +0.15),
    ]
    for r, row in enumerate(sleeves, start=6):
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=val)
            if col == 1:
                label_cell(c)
            elif col == 2:
                style_cell(c)
                c.number_format = "0.000"
            else:
                style_cell(c)
                c.number_format = "0.00%"

    # Section 2: contributions C[i,e] = w[i] · S[i,e]
    ws["A12"] = "Step 2 — Per-sleeve contribution C_{i,e} = w_i × S_{i,e}"
    ws["A12"].font = F_LABEL
    ws["A12"].fill = FILL_PANEL
    ws.merge_cells("A12:G12")

    cols2 = ["sleeve", "", "Defl.", "Infl.", "Stag.", "Refl."]
    for i, h in enumerate(cols2, start=1):
        c = ws.cell(row=13, column=i, value=h)
        header_cell(c)

    # Contribution formulas: C{r,c} = $B{src} * S{src,c}
    for idx in range(5):
        src = 6 + idx          # row of weight + S in step-1 block
        dst = 14 + idx         # row in contribution block
        # Sleeve label = same as step 1
        c = ws.cell(row=dst, column=1, value=f"=A{src}")
        label_cell(c)
        # Empty col 2 (no weight repeat)
        ws.cell(row=dst, column=2, value="")
        for k, src_col in enumerate(["C", "D", "E", "F"], start=3):
            f = f"=$B${src}*{src_col}{src}"
            c = ws.cell(row=dst, column=k, value=f)
            style_cell(c)
            c.number_format = "0.00%"

    # Sum row R^port_e
    sum_row = 19
    c = ws.cell(row=sum_row, column=1, value="R^port_e")
    style_cell(c, font=Font(name="Inter", size=10, bold=True, color=PAL["text_1"]),
               fill=FILL_HEAD, align=ALIGN_LEFT)
    ws.cell(row=sum_row, column=2, value="")
    for k, col_letter in enumerate(["C", "D", "E", "F"], start=3):
        f = f"=SUM({col_letter}14:{col_letter}18)"
        c = ws.cell(row=sum_row, column=k, value=f)
        style_cell(c, font=Font(name="Inter", size=10, bold=True, color=PAL["text_1"]),
                   fill=FILL_HEAD)
        c.number_format = "0.00%"

    # Section 3: comparison vs Table 7.1 byte-exact
    ws["A21"] = "Verification — research/12 §7 Table 7.1 sum row (byte-exact target)"
    ws["A21"].font = F_LABEL
    ws["A21"].fill = FILL_PANEL
    ws.merge_cells("A21:G21")

    targets = [
        ("R^port_e (Table 7.1)", "—", -0.08125, -0.260, -0.0305, +0.11825),
    ]
    for r, row in enumerate(targets, start=22):
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=val)
            if col == 1:
                label_cell(c)
            elif col == 2:
                style_cell(c, align=ALIGN_CENTER)
            else:
                style_cell(c)
                c.number_format = "0.00%"

    # Diff row
    diff_row = 23
    c = ws.cell(row=diff_row, column=1, value="diff (computed − target)")
    label_cell(c)
    ws.cell(row=diff_row, column=2, value="")
    for k, col_letter in enumerate(["C", "D", "E", "F"], start=3):
        f = f"={col_letter}{sum_row}-{col_letter}22"
        c = ws.cell(row=diff_row, column=k, value=f)
        style_cell(c)
        c.number_format = "0.0000%"

    # Footer notes
    notes = [
        "• Shock matrix S verbatim from research/12 §5 Step 1.",
        "• Weights w = Robbins (2014) 30/40/15/7.5/7.5 — Dalio canon via Tony Robbins.",
        "• Sum-row formulas SUMPRODUCT-equivalent via per-cell C_{i,e} + SUM column.",
        "• Diff row should show 0.00% across all four archetypes (computed matches Table 7.1).",
    ]
    for k, n in enumerate(notes, start=25):
        c = ws.cell(row=k, column=1, value=n)
        note_cell(c)
        ws.merge_cells(start_row=k, start_column=1, end_row=k, end_column=6)

    widths = [22, 12, 12, 12, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PAL["bad"][2:]


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main() -> int:
    wb = Workbook()

    # Remove default sheet
    default = wb.active
    wb.remove(default)

    ws_del = wb.create_sheet("4_Deleverage")
    build_deleverage_sheet(ws_del)

    ws_str = wb.create_sheet("12_Stress")
    build_stress_sheet(ws_str)

    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(PILOT_DIR.parent)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
